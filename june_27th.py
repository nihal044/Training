import pandas as pd
import openpyxl
from weasyprint import HTML
# sample data(n)
data=pd.read_csv("employees.csv")
row1=data.sample(n=1)
row1

row2=data.sample(n=5)
row2
# frac data
row3=data.sample(frac=0.25)
if(0.25*(len(data))==len(row3)):
    print(len(data),len(row3))   
#weights
row3
weights=[0.1,0.2,0.3,0.4,0.5]
sampled_data=row2.sample(n=3,weights=weights)
print(sampled_data)
#replace
replace=row2.sample(n=5,replace=True)
replace
#axis
axis=data.sample(n=4,axis=1)
axis
#random_state
random=data.sample(n=5,random_state=1)
random
path=r"E:\Nihal_Rasuri\KTPL\Training\employees.xlsx"
obj=openpyxl.load_workbook(path)
sheet_obj=obj.active

cell_obj=sheet_obj.cell(row=1,column=1)

cell_obj=sheet_obj.max_row

print(cell_obj)

#Simple one to_excel()

data = {
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Edward'],
    'Age': [24, 27, 22, 32, 29],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix']
}

df = pd.DataFrame(data)
sample=df.sample(n=10)
df.to_excel("output.xlsx")

# Using openpyxl to excel

def save_to_excel(data, n=None, frac=None, output_path="data.xlsx"):

    if n is not None:
        sampled_data = data.sample(n=n)
    elif frac is not None:
        sampled_data = data.sample(frac=frac)
    else:
        raise ValueError("enter n or frac value")
    
    sampled_data.to_excel(output_path, index=False, engine='openpyxl')
    print(f"Sampled data saved to {output_path}")

data = {
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Edward'],
    'Age': [24, 27, 22, 32, 29],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix']
}

df = pd.DataFrame(data)
save_to_excel(df, n=2, output_path="sampled_data.xlsx")

#Sample data into html

import pandas as pd

def save_sample_to_html(data, n=None, frac=None, output_path="sampled_data.html"):
 
    if n is not None:
        sampled_data = data.sample(n=n)
    elif frac is not None:
        sampled_data = data.sample(frac=frac)
    else:
        raise ValueError("enter n or frac value")
    
    html_content = sampled_data.to_html()
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Sampled Data</title>
   
    </head>
    <body>
        <h2>Sampled Data</h2>
        {html_content}
    </body>
    </html>
    """

    with open(output_path, "w") as file:
        file.write(html_template)
    
    print(f"Sampled data saved to {output_path}")

data = {
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Edward'],
    'Age': [24, 27, 22, 32, 29],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix']
}

df = pd.DataFrame(data)

save_sample_to_html(df, n=2, output_path="sampled_data.html")

HTML("file:///E:/Nihal_Rasuri/KTPL/Training/sampled_data.html").write_pdf("weasyprint.pdf")
print("saved pdf at location")

#endpoint to generate excel and returning it
from fastapi import FastAPI,Response
from fastapi.responses import FileResponse
import pandas as pd
from weasyprint import HTML

app = FastAPI()

data = {
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Edward'],
    'Age': [24, 27, 22, 32, 29],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix']
}
def generate_excel():
    df = pd.DataFrame(data)
    file_path = "output.xlsx"
    df.to_excel(file_path, index=False)
    return file_path

@app.get("/generate-excel")
def generate_and_return_excel():
    excel_file = generate_excel()
    return FileResponse(excel_file, filename="output.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

#  endpoint Generating pdf and returning it
html_content = """
<!DOCTYPE html>
<html>
<head>
    <title>Sample PDF</title>
</head>
<body>
    <h1>Hello, Fatsapi</h1>
    <p>This is a sample pdf.</p>
</body>
</html>
"""
def generate_pdf():
    pdf_file = "output.pdf"
    HTML(string=html_content).write_pdf(pdf_file)
    return pdf_file

@app.get("/generate_pdf")
def generate_and_return_pdf(response: Response):
    pdf_file = generate_pdf()
    response.headers["Content-Disposition"] = "attachment; filename=output.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return FileResponse(pdf_file, filename="output.pdf", media_type="application/pdf")
